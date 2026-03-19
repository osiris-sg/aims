#!/usr/bin/env python3
"""
AI-powered invoice extraction using Claude API.
Processes all Xero invoices and extracts structured data for each line item.
"""

import json
import os
import time
import re
import anthropic
import openpyxl
from collections import defaultdict

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_PATH = os.path.join(SCRIPT_DIR, 'xero-import-prefilled.json')
DETAIL_FILE = os.path.expanduser("~/Downloads/Receivable_Invoice_Detail (1).xlsx")
SUMMARY_FILE = os.path.expanduser("~/Downloads/Biofuel_Industries_Pte_Ltd_-_Receivable_Invoice_Summary.xlsx")

# Known assets in DB
KNOWN_ASSETS = [
    "AF-100 (AF100)", "AF-30 (AF30)", "AF-40 (AF40)", "AF-60 (AF60)", "AF5 (AF5)", "AF80 (AF80)",
    "Membrane 10-Capacity (MBR-10)", "Membrane 15-Capacity (MBR-15)", "Membrane 20-Capacity (MBR-20)",
    "Membrane 30-Capacity (MBR-30)", "Membrane 40-Capacity (MBR40)", "Membrane 40/50-Capacity (MBR-4050)",
    "Membrane 60-Capacity (MBR-60)",
    "LION-125 (LION125)", "LION-135 (LION135)", "LION-250 (LION250)", "LION-375 (LION375)", "LION-500 (LION500)",
    "Automated Intervention System (AIS)", "Silt Imagery Detection System (SIDS)", "TSS Sensor (TSS)",
    "Submersible Dewatering Pump (KBZ*)",
    "Micro-Grid System (MICROGRID)", "ECM Plant (ECMPLANT)", "Generator 125KVA (GEN125)",
    "Denyo Soundproof Generator (DENYOGEN)", "APF-10 (APF10)", "APF-30 (APF30)", "APF-40 (APF40)",
    "APF-80 (APF80)", "APF-90 (APF90)", "APF-100 (APF100)",
    "Solar Panel (SOLARPANEL)", "FRP Holding Tank (HOLDINGTANK)",
    "Tipper Lorry (TIPPERLORRY)", "Excavator (EXCAVATOR)",
    "Disposal Service (SVC-DISPOSAL)", "Transport Service (SVC-TRANSPORT)",
    "Installation Service (SVC-INSTALL)", "Desilting Service (SVC-DESILT)",
    "Supply of Materials (SVC-SUPPLY)", "Manpower / Labour (SVC-MANPOWER)",
    "Credit Note Adjustment (SVC-CN)", "Balance Brought Forward (SVC-BF)",
]

client = anthropic.Anthropic()

SYSTEM_PROMPT = """You are a data extraction assistant for Biofuel Industries Pte Ltd, a company that rents and sells water treatment equipment (MBR systems, APF systems, AF systems), generators, micro-grid systems, pumps, and provides disposal/transport services.

Your job is to extract structured data from invoice line items. For each invoice, you'll receive the full reference text and all line items with their descriptions.

Known assets in the database:
""" + "\n".join(f"- {a}" for a in KNOWN_ASSETS) + """

Rules:
1. Each line item with qty > 0 is a PRODUCT line. Extract: asset_name, sku_key, serial_numbers (array), category, uom
2. Lines with qty = 0 are usually REFERENCE lines containing: project name, location, site office, DO number/date, contact info, PO/quotation refs
3. "MBR 01 at Tengah C6" means MBR unit #1 deployed at Tengah C6 - you need to check the full description for the MBR model/capacity and serial number
4. Serial numbers appear as "S/No.", "s/n", "S/NO.", or patterns like "MG20250079", "2021051069", "AF40 0007"
5. If qty > 1, there may be multiple serial numbers in the description
6. The full_reference field often contains the asset model and serial (e.g. "BI202509041 (2nd mth - Canberra DO/BT202508-001 1xLION375 MG20250079)")
7. For rental items, status should be "rental". For sold items, "sold".
8. Extract project name from "Project:" lines, location from "Location:" lines, DO date from "Our DO No. ... dated DD/MM/YYYY"

Return a JSON object (no markdown, just raw JSON) with this structure:
{
  "product_lines": [
    {
      "line_index": 0,
      "asset_name": "LION-375",
      "sku_key": "LION375",
      "serial_numbers": ["MG20250079"],
      "category": "Equipment",
      "uom": "UNIT",
      "is_reference_line": false
    }
  ],
  "reference_lines": [0, 3, 4],
  "project_name": "Canberra Crescent Residences",
  "site_office_name": "51 Canberra Cres",
  "site_office_address": "51 Canberra Cres, Singapore 752106",
  "do_date": "2025-08-16",
  "do_number": "DO/BT202508-001",
  "contact_name": "Mr Feng Tianru",
  "contact_phone": "9649 6903"
}
"""


def load_summary():
    """Load customer mapping from summary file."""
    wb = openpyxl.load_workbook(SUMMARY_FILE, data_only=True)
    ws = wb['Receivable Invoice Summary']

    invoice_to_customer = {}
    invoice_summary = {}
    current_customer = None

    for row in ws.iter_rows(min_row=5, values_only=False):
        col_a = row[0].value
        col_b = row[1].value
        col_h = row[7].value

        if col_a is None:
            continue
        col_a_str = str(col_a).strip()
        if col_a_str.startswith("Total"):
            continue
        if col_b is None and col_h is None:
            current_customer = col_a_str
            continue
        if col_b is not None and current_customer:
            invoice_to_customer[col_a_str] = current_customer
            invoice_summary[col_a_str] = {
                "gross": float(row[4].value) if row[4].value else 0,
                "balance": float(row[5].value) if row[5].value else 0,
                "status": str(row[6].value) if row[6].value else None,
                "source": str(row[7].value) if row[7].value else None,
            }

    wb.close()
    return invoice_to_customer, invoice_summary


def load_detail():
    """Load all invoices with full descriptions from detail file."""
    wb = openpyxl.load_workbook(DETAIL_FILE, data_only=True)
    ws = wb['Receivable Invoice Detail']

    invoices = {}
    current_invoice = None

    for row in ws.iter_rows(min_row=5, values_only=True):
        date, source, ref, item_code, desc, qty, unit_price, discount, tax, gross, inv_total, status = row

        if date is None and source is None and ref is None:
            continue

        col_a_str = str(date).strip() if date else ''

        if col_a_str.startswith('Total '):
            current_invoice = None
            continue

        if source is None and desc is None and col_a_str:
            current_invoice = col_a_str
            if current_invoice not in invoices:
                invoices[current_invoice] = []
            continue

        if current_invoice and source:
            invoices[current_invoice].append({
                'date': date.strftime('%Y-%m-%d') if hasattr(date, 'strftime') else str(date) if date else '',
                'source': str(source) if source else '',
                'reference': str(ref) if ref else '',
                'description': str(desc) if desc else '',
                'quantity': float(qty) if qty else 0,
                'unit_price': float(unit_price) if unit_price else 0,
                'discount': float(discount) if discount else 0,
                'tax': float(tax) if tax else 0,
                'gross': float(gross) if gross else 0,
                'invoice_total': float(inv_total) if inv_total else 0,
                'status': str(status) if status else '',
            })

    wb.close()
    return invoices


def extract_with_ai(inv_number, items, reference):
    """Send invoice to Claude for extraction."""
    user_msg = f"Invoice: {inv_number}\nFull Reference: {reference}\n\nLine items:\n"
    for i, item in enumerate(items):
        user_msg += f"\n[{i}] Qty: {item['quantity']} | Unit: ${item['unit_price']} | Gross: ${item['gross']}\n"
        user_msg += f"    Description: {item['description'][:500]}\n"

    try:
        response = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=2000,
            system=SYSTEM_PROMPT,
            messages=[{"role": "user", "content": user_msg}],
        )

        text = response.content[0].text.strip()
        # Remove markdown code fences if present
        text = re.sub(r'^```json\s*', '', text)
        text = re.sub(r'\s*```$', '', text)
        text = text.strip()
        # Try to parse JSON from response
        if text.startswith('{'):
            return json.loads(text)
        # Try to find JSON in the response
        match = re.search(r'\{[\s\S]*\}', text)
        if match:
            return json.loads(match.group())
        print(f"  WARNING: Could not parse response for {inv_number}")
        return None
    except Exception as e:
        print(f"  ERROR for {inv_number}: {e}")
        return None


def main():
    print("Loading Excel files...")
    invoice_to_customer, invoice_summary = load_summary()
    detail_invoices = load_detail()

    print(f"Summary: {len(invoice_to_customer)} invoices with customers")
    print(f"Detail: {len(detail_invoices)} invoices with line items")

    # Filter out voided/deleted
    skip_statuses = {'Voided', 'Deleted'}

    all_invoices = []
    to_process = []

    for inv_number, items in detail_invoices.items():
        if not items:
            continue
        status = items[0].get('status', '')
        if status in skip_statuses:
            continue

        customer = invoice_to_customer.get(inv_number)
        summary = invoice_summary.get(inv_number, {})
        reference = items[0].get('reference', '')
        date = items[0].get('date', '')
        source = items[0].get('source', '')

        inv_data = {
            'invoice_number': inv_number,
            'date': date,
            'customer': customer or 'Unknown',
            'customer_matched': customer is not None,
            'status': summary.get('status', status),
            'source': summary.get('source', source),
            'gross': summary.get('gross', items[0].get('invoice_total', 0)),
            'balance': summary.get('balance', 0),
            'reference': reference,
            'raw_items': items,
        }
        to_process.append(inv_data)

    print(f"\nInvoices to process with AI: {len(to_process)}")

    # Process in batches
    processed = 0
    errors = 0

    for inv_data in to_process:
        inv_number = inv_data['invoice_number']
        items = inv_data['raw_items']
        reference = inv_data['reference']

        processed += 1
        if processed % 50 == 0:
            print(f"  Progress: {processed}/{len(to_process)}")

        # Call AI
        ai_result = extract_with_ai(inv_number, items, reference)

        if not ai_result:
            errors += 1
            # Fallback: basic extraction
            ai_result = {
                'product_lines': [],
                'reference_lines': [],
                'project_name': None,
                'site_office_name': None,
                'site_office_address': None,
                'do_date': None,
                'contact_name': None,
                'contact_phone': None,
            }

        # Build line items with AI results
        product_map = {}
        for pl in ai_result.get('product_lines', []):
            product_map[pl.get('line_index', -1)] = pl

        ref_lines = set(ai_result.get('reference_lines', []))

        line_items = []
        for i, item in enumerate(items):
            is_ref = i in ref_lines or item['quantity'] == 0

            if i in product_map:
                pl = product_map[i]
                line_items.append({
                    'description': item['description'],
                    'quantity': item['quantity'],
                    'unit_price': item['unit_price'],
                    'discount': item['discount'],
                    'tax': item['tax'],
                    'gross': item['gross'],
                    'is_reference_line': False,
                    'asset_match': {
                        'key': pl.get('sku_key', ''),
                        'name': pl.get('asset_name', ''),
                        'sku': pl.get('sku_key', ''),
                        'category': pl.get('category', ''),
                        'exists_in_db': False,  # Will be resolved by frontend
                        'needs_sku': not pl.get('sku_key'),
                    },
                    'confidence': 'high' if pl.get('sku_key') else 'medium',
                    'match_reason': 'AI extraction',
                    'location': None,
                    'serial_numbers': pl.get('serial_numbers', []),
                })
            elif is_ref:
                line_items.append({
                    'description': item['description'],
                    'quantity': item['quantity'],
                    'unit_price': item['unit_price'],
                    'discount': item['discount'],
                    'tax': item['tax'],
                    'gross': item['gross'],
                    'is_reference_line': True,
                    'asset_match': None,
                    'confidence': None,
                    'match_reason': 'Reference line',
                    'location': None,
                    'serial_numbers': [],
                })
            else:
                line_items.append({
                    'description': item['description'],
                    'quantity': item['quantity'],
                    'unit_price': item['unit_price'],
                    'discount': item['discount'],
                    'tax': item['tax'],
                    'gross': item['gross'],
                    'is_reference_line': False,
                    'asset_match': None,
                    'confidence': None,
                    'match_reason': 'AI could not determine',
                    'location': None,
                    'serial_numbers': [],
                })

        inv_data['line_items'] = line_items
        inv_data['project_name'] = ai_result.get('project_name')
        inv_data['project_location'] = ai_result.get('site_office_address') or ai_result.get('project_name')
        inv_data['site_office_name'] = ai_result.get('site_office_name')
        inv_data['site_office_address'] = ai_result.get('site_office_address')
        inv_data['do_date'] = ai_result.get('do_date')
        inv_data['do_number'] = ai_result.get('do_number')
        inv_data['contact_name'] = ai_result.get('contact_name')
        inv_data['contact_phone'] = ai_result.get('contact_phone')
        inv_data['review_status'] = 'pending'

        # Remove raw_items
        del inv_data['raw_items']
        del inv_data['reference']

        all_invoices.append(inv_data)

        # Rate limiting - small delay
        time.sleep(0.3)

    # Sort by date
    all_invoices.sort(key=lambda x: x.get('date') or '')

    # Build summary
    total_items = sum(len(inv['line_items']) for inv in all_invoices)
    ref_items = sum(1 for inv in all_invoices for li in inv['line_items'] if li['is_reference_line'])
    matched = sum(1 for inv in all_invoices for li in inv['line_items'] if li.get('asset_match') and li['asset_match'].get('sku'))

    output = {
        'summary': {
            'total_invoices': len(all_invoices),
            'total_line_items': total_items,
            'reference_lines': ref_items,
            'product_lines': total_items - ref_items,
            'matched_high_confidence': matched,
            'matched_medium_confidence': 0,
            'matched_low_confidence': 0,
            'unmatched': total_items - ref_items - matched,
            'match_rate': f"{matched / max(1, total_items - ref_items) * 100:.1f}%",
            'auto_confirmable': f"{matched / max(1, total_items - ref_items) * 100:.1f}%",
        },
        'invoices': all_invoices,
    }

    print(f"\nWriting to {OUTPUT_PATH}...")
    with open(OUTPUT_PATH, 'w') as f:
        json.dump(output, f, indent=2, default=str)

    print(f"\nDone!")
    print(f"  Total invoices: {len(all_invoices)}")
    print(f"  Total line items: {total_items}")
    print(f"  Reference lines: {ref_items}")
    print(f"  Matched products: {matched}")
    print(f"  Errors: {errors}")


if __name__ == '__main__':
    main()
