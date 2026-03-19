#!/usr/bin/env python3
"""
Analyze Xero Excel exports and produce a JSON report for review
before importing into the AIMS database.

Input files:
  1. Receivable Invoice Detail - line item details
  2. Receivable Invoice Summary - invoice-level data with customer names

Output:
  xero-import-report.json - structured report of invoices, assets, and customers
"""

import json
import os
import re
from collections import defaultdict
from datetime import datetime

import openpyxl

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

DETAIL_FILE = os.path.expanduser(
    "~/Downloads/Receivable_Invoice_Detail (1).xlsx"
)
DETAIL_SHEET = "Receivable Invoice Detail"

SUMMARY_FILE = os.path.expanduser(
    "~/Downloads/Biofuel_Industries_Pte_Ltd_-_Receivable_Invoice_Summary.xlsx"
)
SUMMARY_SHEET = "Receivable Invoice Summary"

OUTPUT_FILE = os.path.join(os.path.dirname(__file__), "xero-import-report.json")

ORGANIZATION_ID = "52e90ba8-bfbd-48b0-bb76-4f9667bf74f1"

# Known assets in the database: (name, skuKey)
KNOWN_ASSETS = [
    ("AF-100", "AF100"),
    ("AF-30", "AF30"),
    ("AF-40", "AF40"),
    ("AF-60", "AF60"),
    ("AF5", "AF5"),
    ("AF80", "AF80"),
    ("Membrane 10-Capacity", "MBR-10"),
    ("Membrane 10-Capacity", "MBR10"),
    ("Membrane 15-Capacity", "MBR-15"),
    ("Membrane 20-Capacity", "MBR-20"),
    ("Membrane 30-Capacity", "MBR30"),
    ("Membrane 30-Capacity", "MBR-30"),
    ("Membrane 40-Capacity", "MBR40"),
    ("Membrane 40/50-Capacity", "MBR-4050"),
    ("Membrane 60-Capacity", "MBR-60"),
    ("LION-125", "LION-125"),
    ("LION-135", "LION-135"),
    ("LION-250", "LION-250"),
    ("LION-375", "LION-375"),
    ("LION-500", "LION-500"),
    ("Automated Intervention System", "AIS"),
    ("Silt Imagery Detection System", "SIDS"),
    ("TSS Sensor", "TSS"),
    # KBZ submersible dewatering pump variants
    ("Submersible Dewatering Pump", "KBZ"),
]

# Patterns that indicate a reference / metadata line, not a product
REFERENCE_LINE_PATTERNS = [
    r"(?i)^PO\s*No",
    r"(?i)^DO\s*No",
    r"(?i)^Qtn\s*Ref",
    r"(?i)^Location\s*:",
    r"(?i)^Project\s*:",
    r"(?i)^Attn\s*:",
    r"(?i)^Remarks?\s*:",
    r"(?i)^Our\s+DO",
    r"(?i)^Your\s+PO",
    r"(?i)^Service\s+Order",
    r"(?i)^Rental\s+[Pp]eriod",
    r"(?i)^Duration",
    r"(?i)^Pro-rat",
    r"(?i)^Discount",
    r"(?i)^Before\s+discount",
    r"(?i)^Further\s+[Dd]iscount",
    r"(?i)^Off-[Hh]ired",
]

# ---------------------------------------------------------------------------
# Asset matching rules
# ---------------------------------------------------------------------------

def build_asset_matchers():
    """
    Return a list of (compiled_regex, canonical_name, exists_in_db) tuples.
    Order matters -- more specific patterns first.
    """
    matchers = []

    def add(pattern, canonical, exists):
        matchers.append((re.compile(pattern, re.IGNORECASE), canonical, exists))

    # AF variants (order: longer model numbers first)
    add(r"AF[\s-]?100", "AF-100", True)
    add(r"AF[\s-]?80", "AF80", True)
    add(r"AF[\s-]?60", "AF60", True)
    add(r"AF[\s-]?40", "AF-40", True)
    add(r"AF[\s-]?30", "AF-30", True)
    add(r"AF[\s-]?5(?!\d)", "AF5", True)

    # APF variants (not in DB -- need new assets)
    add(r"APF[\s-]?100", "APF-100", False)
    add(r"APF[\s-]?90", "APF-90", False)
    add(r"APF[\s-]?80", "APF-80", False)
    add(r"APF[\s-]?40", "APF-40", False)
    add(r"APF[\s-]?30", "APF-30", False)
    add(r"APF[\s-]?10", "APF-10", False)

    # MBR / Membrane variants
    add(r"MBR[\s-]?60", "MBR-60", True)
    add(r"MBR[\s-]?40/?50", "MBR-4050", True)
    add(r"MBR[\s-]?40", "MBR40", True)
    add(r"MBR[\s-]?30", "MBR-30", True)
    add(r"MBR[\s-]?20", "MBR-20", True)
    add(r"MBR[\s-]?15", "MBR-15", True)
    add(r"MBR[\s-]?10", "MBR-10", True)

    # ECM Plant
    add(r"ECM[\s-]?\d*\s*(?:Plant|System|system)", "ECM Plant", False)
    add(r"ECM[\s-]?\d+", "ECM Plant", False)
    add(r"\bECM\b", "ECM Plant", False)

    # LION variants
    add(r"LION[\s-]?500", "LION-500", True)
    add(r"LION[\s-]?375", "LION-375", True)
    add(r"LION[\s-]?250", "LION-250", True)
    add(r"LION[\s-]?135", "LION-135", True)
    add(r"LION[\s-]?125", "LION-125", True)

    # SIDS
    add(r"\bSIDS\b", "SIDS", True)
    add(r"Silt\s+Imager", "SIDS", True)

    # TSS / CCTV monitoring
    add(r"\bTSS\b", "TSS", True)
    add(r"\bCCTV\b", "TSS", True)

    # AIS
    add(r"\bAIS\b", "AIS", True)

    # KBZ pumps
    add(r"\bKBZ", "KBZ Submersible Pump", True)

    # Generators
    add(r"\b\d+\s*KVA\b", "Generator", False)
    add(r"\bGenerator\b", "Generator", False)
    add(r"\bGenset\b", "Generator", False)

    # Micro-Grid
    add(r"Micro[\s-]?Grid", "Micro-Grid", False)

    # Submersible pump (generic, not KBZ)
    add(r"[Ss]ubmersible\s+(?:[Dd]ewatering\s+)?[Pp]ump", "Submersible Pump", False)

    # Weighbridge
    add(r"[Ww]eighbridge", "Weighbridge", False)

    # Trailer
    add(r"\b[Tt]railer\b", "Trailer", False)

    # Solar Panel (standalone, not part of SIDS description already matched)
    add(r"[Ss]olar\s+[Pp]anel", "Solar Panel", False)

    return matchers


ASSET_MATCHERS = build_asset_matchers()

# Patterns for vehicle/DO delivery lines (number plate + date pattern)
VEHICLE_DO_PATTERN = re.compile(
    r"^\d{4,5}\s*-?\s*X[A-Z]\d{3,4}[A-Z]?\s+on\s+", re.IGNORECASE
)
# DO reference lines like "DO 3383  on 02/07/2021"
DO_LINE_PATTERN = re.compile(r"^DO\s+\d+\s+on\s+", re.IGNORECASE)
# Trip count lines like "02/02/2026 - 38 trips"
TRIP_COUNT_PATTERN = re.compile(r"^\d{2}/\d{2}/\d{4}\s*-\s*\d+\s+trips?", re.IGNORECASE)


def is_reference_line(description: str) -> bool:
    """Check if a line item description is just a reference/metadata line."""
    desc = description.strip()
    if not desc:
        return True

    # Check explicit reference patterns
    for pattern in REFERENCE_LINE_PATTERNS:
        if re.search(pattern, desc):
            return True

    return False


def is_delivery_or_transport_line(description: str) -> bool:
    """Check if a line is about vehicle delivery / DO trips."""
    desc = description.strip()
    if VEHICLE_DO_PATTERN.match(desc):
        return True
    if DO_LINE_PATTERN.match(desc):
        return True
    if TRIP_COUNT_PATTERN.match(desc):
        return True
    return False


def match_asset(description: str):
    """
    Try to match a description to a known asset.
    Returns (canonical_name, exists_in_db) or (None, None).
    """
    if not description:
        return None, None

    for regex, canonical, exists in ASSET_MATCHERS:
        if regex.search(description):
            return canonical, exists

    return None, None


def extract_locations(description: str) -> list:
    """Extract location info from description text."""
    locations = []
    # "Location:" pattern
    loc_matches = re.findall(r"[Ll]ocation\s*:\s*([^\n]+)", description)
    for m in loc_matches:
        loc = m.strip().rstrip(".")
        if loc:
            locations.append(loc)
    return locations


def categorize_unmatched(description: str) -> str:
    """Attempt to categorize an unmatched line item."""
    desc = description.lower()

    if "disposal" in desc:
        return "Disposal Service"
    if "transport" in desc:
        return "Transport Service"
    if "supply" in desc:
        return "Supply"
    if any(w in desc for w in ["pump", "valve", "e-valve", "backwash"]):
        return "Pump/Parts"
    if any(w in desc for w in ["rental", "hire"]):
        return "Equipment Rental"
    if any(w in desc for w in ["sale", "sold"]):
        return "Sales"
    if any(w in desc for w in ["repair", "service", "maintenance", "calibration"]):
        return "Service/Maintenance"
    if "piping" in desc or "pipe" in desc:
        return "Piping Works"
    if "installation" in desc or "installn" in desc or "instaln" in desc:
        return "Installation"
    if any(w in desc for w in ["concrete", "earth", "soil", "hardcore", "rca"]):
        return "Material/Disposal"
    if "b/f" in desc or "b/f" in desc.replace(" ", ""):
        return "Balance Brought Forward"
    if "credit" in desc:
        return "Credit"
    if is_delivery_or_transport_line(description):
        return "Delivery/Transport"

    return "Other"


# ---------------------------------------------------------------------------
# Parsing
# ---------------------------------------------------------------------------

def parse_summary_file(filepath: str, sheet_name: str) -> dict:
    """
    Parse the summary file and build:
      - invoice_to_customer: {invoice_number: customer_name}
      - invoice_summary: {invoice_number: {date, gross, balance, status, source}}
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[sheet_name]

    invoice_to_customer = {}
    invoice_summary = {}
    current_customer = None

    for row in ws.iter_rows(min_row=5, values_only=False):  # skip title rows
        col_a = row[0].value
        col_b = row[1].value  # Invoice Date
        col_h = row[7].value  # Source

        if col_a is None:
            continue

        col_a_str = str(col_a).strip()

        # Skip total rows and empty
        if col_a_str.startswith("Total"):
            continue

        # Customer header: has value in col A but nothing in col B and col H
        if col_b is None and col_h is None:
            current_customer = col_a_str
            continue

        # Invoice row
        if col_b is not None and current_customer:
            invoice_number = col_a_str
            invoice_to_customer[invoice_number] = current_customer
            invoice_summary[invoice_number] = {
                "date": col_b.strftime("%Y-%m-%d") if isinstance(col_b, datetime) else str(col_b),
                "expected_date": (
                    row[2].value.strftime("%Y-%m-%d")
                    if isinstance(row[2].value, datetime)
                    else str(row[2].value) if row[2].value else None
                ),
                "reference": str(row[3].value) if row[3].value else None,
                "gross": float(row[4].value) if row[4].value else 0,
                "balance": float(row[5].value) if row[5].value else 0,
                "status": str(row[6].value) if row[6].value else None,
                "source": str(row[7].value) if row[7].value else None,
                "invoice_sent": str(row[8].value) if row[8].value else None,
            }

    wb.close()
    return invoice_to_customer, invoice_summary


def parse_detail_file(filepath: str, sheet_name: str) -> dict:
    """
    Parse the detail file. Structure:
      - Invoice number header row: col A has invoice number, nothing else meaningful
      - Line item rows: all columns populated
      - Total row: col A starts with "Total <invoice_number>"

    Returns: {invoice_number: [list of line item dicts]}
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[sheet_name]

    invoices = {}
    current_invoice = None
    header_row = 4  # Row 4 has column headers

    for row in ws.iter_rows(min_row=header_row + 1, values_only=False):
        col_a = row[0].value
        col_b = row[1].value  # Source
        col_e = row[4].value  # Description

        if col_a is None:
            continue

        col_a_str = str(col_a).strip()

        # Skip total rows
        if col_a_str.startswith("Total "):
            current_invoice = None
            continue

        # If col_b (Source) is None and col_e (Description) is None,
        # this is an invoice header row
        if col_b is None and col_e is None:
            current_invoice = col_a_str
            if current_invoice not in invoices:
                invoices[current_invoice] = []
            continue

        # Line item row
        if current_invoice and col_b is not None:
            date_val = col_a
            line_item = {
                "date": (
                    date_val.strftime("%Y-%m-%d")
                    if isinstance(date_val, datetime)
                    else str(date_val) if date_val else None
                ),
                "source": str(col_b) if col_b else None,
                "reference": str(row[2].value) if row[2].value else None,
                "item_code": str(row[3].value) if row[3].value else None,
                "description": str(row[4].value) if row[4].value else "",
                "quantity": float(row[5].value) if row[5].value else 0,
                "unit_price": float(row[6].value) if row[6].value else 0,
                "discount": float(row[7].value) if row[7].value else 0,
                "tax": float(row[8].value) if row[8].value else 0,
                "gross": float(row[9].value) if row[9].value else 0,
                "invoice_total": float(row[10].value) if row[10].value else 0,
                "status": str(row[11].value) if row[11].value else None,
            }
            invoices[current_invoice].append(line_item)

    wb.close()
    return invoices


# ---------------------------------------------------------------------------
# Analysis
# ---------------------------------------------------------------------------

def analyze(invoice_to_customer, invoice_summary, detail_invoices):
    """Build the full analysis report."""

    skipped = {"voided": 0, "deleted": 0, "draft": 0}
    matched_assets_count = 0
    unmatched_assets_count = 0
    total_line_items = 0
    customers_seen = set()
    locations_seen = defaultdict(lambda: {"count": 0, "customers": set()})
    new_assets_tracker = defaultdict(lambda: {"count": 0, "descriptions": [], "category": ""})

    output_invoices = []

    for inv_number, line_items in detail_invoices.items():
        if not line_items:
            continue

        # Get status from first line item
        status = line_items[0].get("status", "")
        source = line_items[0].get("source", "")
        date = line_items[0].get("date", "")

        # Count skipped statuses
        if status == "Voided":
            skipped["voided"] += 1
            continue
        if status == "Deleted":
            skipped["deleted"] += 1
            continue
        if status == "Draft":
            skipped["draft"] += 1
            # Don't skip drafts, just flag them

        # Customer lookup
        customer = invoice_to_customer.get(inv_number, None)
        customer_matched = customer is not None
        if customer:
            customers_seen.add(customer)

        # Summary data
        summary = invoice_summary.get(inv_number, {})
        gross_total = summary.get("gross", None)
        balance = summary.get("balance", None)

        # If no summary data, use detail data
        if gross_total is None and line_items:
            gross_total = line_items[0].get("invoice_total", 0)
        if balance is None:
            balance = 0

        # Process line items
        processed_items = []
        for item in line_items:
            total_line_items += 1
            desc = item["description"]
            first_line = desc.split("\n")[0].strip() if desc else ""

            # Check if it is a reference line
            ref_line = is_reference_line(first_line)

            # Check if it is a delivery/transport line
            if not ref_line and is_delivery_or_transport_line(first_line):
                ref_line = False  # These are billable lines, not references

            # Try to match asset from full description
            asset_name, asset_exists = match_asset(desc)

            if asset_name and not ref_line:
                matched_assets_count += 1
            elif not ref_line and not asset_name:
                unmatched_assets_count += 1
                # Track potential new assets
                category = categorize_unmatched(desc)
                if category not in ("Balance Brought Forward", "Other", "Credit"):
                    key = category
                    # Try to be more specific
                    if "disposal" in desc.lower():
                        key = "Disposal Service"
                    elif is_delivery_or_transport_line(desc):
                        key = "Delivery/Transport"
                    new_assets_tracker[key]["count"] += 1
                    if len(new_assets_tracker[key]["descriptions"]) < 3:
                        new_assets_tracker[key]["descriptions"].append(
                            first_line[:150]
                        )
                    new_assets_tracker[key]["category"] = category

            # Extract locations
            locs = extract_locations(desc)
            for loc in locs:
                locations_seen[loc]["count"] += 1
                if customer:
                    locations_seen[loc]["customers"].add(customer)

            processed_items.append({
                "description": desc,
                "quantity": item["quantity"],
                "unit_price": item["unit_price"],
                "discount": item["discount"],
                "tax": item["tax"],
                "gross": item["gross"],
                "matched_asset": asset_name,
                "asset_exists": asset_exists if asset_name else None,
                "is_reference_line": ref_line,
                "item_code": item["item_code"],
                "reference": item.get("reference"),
            })

        inv_record = {
            "invoice_number": inv_number,
            "date": date,
            "customer": customer,
            "customer_matched": customer_matched,
            "status": status,
            "source": source,
            "gross": gross_total,
            "balance": balance,
            "is_draft": status == "Draft",
            "line_items": processed_items,
        }
        output_invoices.append(inv_record)

    # Sort invoices by date
    output_invoices.sort(key=lambda x: x.get("date") or "")

    # Build new_assets_needed from the tracker + unmatched asset names
    # Also collect new assets that were identified by the matcher but don't exist
    new_asset_names = defaultdict(lambda: {"count": 0, "descriptions": []})
    for inv in output_invoices:
        for item in inv["line_items"]:
            if item["matched_asset"] and item["asset_exists"] is False:
                name = item["matched_asset"]
                new_asset_names[name]["count"] += 1
                desc = item["description"]
                if (
                    len(new_asset_names[name]["descriptions"]) < 3
                    and desc not in new_asset_names[name]["descriptions"]
                ):
                    new_asset_names[name]["descriptions"].append(desc[:150])

    new_assets_needed = []

    # Known product patterns that need new assets
    for name, data in sorted(new_asset_names.items()):
        sku = re.sub(r"[\s-]", "", name).upper()
        category = "Equipment"
        if "pump" in name.lower():
            category = "Pump"
        elif "generator" in name.lower() or "genset" in name.lower():
            category = "Generator"
        elif "grid" in name.lower():
            category = "Power System"
        elif "panel" in name.lower():
            category = "Accessory"
        elif "weighbridge" in name.lower():
            category = "Equipment"
        elif "trailer" in name.lower():
            category = "Vehicle"

        new_assets_needed.append({
            "suggested_name": name,
            "suggested_sku": sku,
            "category": category,
            "occurrences": data["count"],
            "sample_descriptions": data["descriptions"],
        })

    # Service-type new assets from unmatched
    for key, data in sorted(new_assets_tracker.items()):
        new_assets_needed.append({
            "suggested_name": key,
            "suggested_sku": re.sub(r"[\s/]", "-", key).upper(),
            "category": data["category"],
            "occurrences": data["count"],
            "sample_descriptions": data["descriptions"],
        })

    # New projects from locations
    new_projects = []
    for loc, data in sorted(locations_seen.items(), key=lambda x: -x[1]["count"]):
        new_projects.append({
            "location": loc,
            "occurrences": data["count"],
            "customers": sorted(data["customers"]),
        })

    report = {
        "organization_id": ORGANIZATION_ID,
        "generated_at": datetime.now().isoformat(),
        "summary": {
            "total_invoices": len(output_invoices),
            "total_line_items": total_line_items,
            "matched_assets": matched_assets_count,
            "unmatched_assets": unmatched_assets_count,
            "matched_customers": len(customers_seen),
            "new_assets_needed": len(new_assets_needed),
            "new_projects_needed": len(new_projects),
            "skipped_invoices": skipped,
        },
        "new_assets_needed": new_assets_needed,
        "new_projects_needed": new_projects,
        "invoices": output_invoices,
    }

    return report


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("Parsing summary file...")
    invoice_to_customer, invoice_summary = parse_summary_file(
        SUMMARY_FILE, SUMMARY_SHEET
    )
    print(f"  Found {len(invoice_to_customer)} invoices with customer mapping")
    print(f"  Found {len(set(invoice_to_customer.values()))} unique customers")

    print("Parsing detail file...")
    detail_invoices = parse_detail_file(DETAIL_FILE, DETAIL_SHEET)
    print(f"  Found {len(detail_invoices)} invoices with line items")

    print("Analyzing...")
    report = analyze(invoice_to_customer, invoice_summary, detail_invoices)

    print(f"\n--- Summary ---")
    s = report["summary"]
    print(f"  Total invoices (after skipping voided/deleted): {s['total_invoices']}")
    print(f"  Total line items: {s['total_line_items']}")
    print(f"  Matched to existing assets: {s['matched_assets']}")
    print(f"  Unmatched (need review): {s['unmatched_assets']}")
    print(f"  Unique customers found: {s['matched_customers']}")
    print(f"  Skipped: voided={s['skipped_invoices']['voided']}, "
          f"deleted={s['skipped_invoices']['deleted']}, "
          f"draft={s['skipped_invoices']['draft']}")
    print(f"  New assets suggested: {s['new_assets_needed']}")
    print(f"  New projects/locations: {s['new_projects_needed']}")

    if report["new_assets_needed"]:
        print(f"\n--- New Assets Needed ---")
        for asset in report["new_assets_needed"]:
            print(f"  {asset['suggested_name']} ({asset['suggested_sku']}) "
                  f"- {asset['occurrences']} occurrences [{asset['category']}]")

    if report["new_projects_needed"]:
        print(f"\n--- Locations/Projects Found ---")
        for proj in report["new_projects_needed"][:10]:
            print(f"  {proj['location']} - {proj['occurrences']} occurrences "
                  f"({', '.join(proj['customers'][:3])})")

    print(f"\nWriting report to: {OUTPUT_FILE}")
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(report, f, indent=2, ensure_ascii=False, default=str)

    print("Done.")


if __name__ == "__main__":
    main()
